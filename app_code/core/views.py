
from django.http import HttpResponse
from PIL import Image

from django.shortcuts import get_object_or_404, render, redirect
import openpyxl
from . models import CodigoPulseira, Pulseira


# Create your views here.
import openpyxl
from django.shortcuts import render
from .models import CodigoPulseira

import openpyxl
from django.shortcuts import render
from .models import CodigoPulseira

def importacao(request):
    if request.method == "GET":
        return render(request, 'core/importacao.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # List the names of worksheets
        sheet_names = wb.sheetnames
        print(sheet_names)  # Print the names of all worksheets

        # Choose the correct worksheet by name
        worksheet = wb[sheet_names[0]]  # Choose the first worksheet, you may adjust this based on your file

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        for row in excel_data:
            codigo = row[0]  # Assuming the code is in the first column
            codigo_pulseira, created = CodigoPulseira.objects.get_or_create(codigo=codigo)
            if created:
                codigo_pulseira.save()

        return render(request, 'core/importacao.html', {"mensagem": "Códigos importados com sucesso!"})
    
def gerar_imagem_composta(qr_codes, largura_coluna, altura, margem):
    total_qr_codes = len(qr_codes)
    colunas = 3
    linhas = (total_qr_codes + colunas - 1) // colunas

    largura = colunas * (largura_coluna + margem) - margem
    imagem_composta = Image.new("RGB", (largura, altura), "white")

    for i, qr_code in enumerate(qr_codes):
        coluna = i % colunas
        linha = i // colunas
        x = coluna * (largura_coluna + margem)
        y = linha * altura

        imagem_qr_code = Image.open(qr_code.qrcode.path)
        imagem_composta.paste(imagem_qr_code, (x, y))

    return imagem_composta

def imprimir_pulseiras(request):
    qr_codes = CodigoPulseira.objects.all()  # Obtém todos os códigos de pulseira com seus QR codes

    # Gere a imagem composta usando a função gerar_imagem_composta
    largura_coluna = 33  # Em milímetros
    altura = 21  # Em milímetros
    margem = 5  # Em milímetros
    imagem_composta = gerar_imagem_composta(qr_codes, largura_coluna, altura, margem)

    # Crie uma resposta HttpResponse contendo a imagem
    response = HttpResponse(content_type="image/png")
    imagem_composta.save(response, "PNG")

    return response

def home(request):
    mensagem = None

    if request.method == 'POST':
        codigo = request.POST.get('codigo')
        try:
            codigo_pulseira = CodigoPulseira.objects.get(codigo=codigo, status=False)
            codigo_pulseira.status = True  # Marca o código como usado (Liberado)
            codigo_pulseira.save()
            mensagem = 'Liberado'
        except CodigoPulseira.DoesNotExist:
            mensagem = 'Bloqueado'

    context = {'mensagem': mensagem}
    return render(request, 'core/index.html', context)


def imprimir_pulseiras(request):
    if request.method == "POST":
        ids = request.POST.getlist("pulseiras")  # Lista de IDs selecionados
        pulseiras = Pulseira.objects.filter(id__in=ids)

        return render(request, 'core/imprimir_pulseiras.html', {'pulseiras': pulseiras})
