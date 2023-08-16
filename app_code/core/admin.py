from django.contrib import admin
import openpyxl
from django import forms
from .models import CodigoPulseira, Pulseira
import qrcode
from django.core.files.base import ContentFile

# Register your models here.

admin.site.register(Pulseira)





@admin.register(CodigoPulseira)
class CodigoPulseiraAdmin(admin.ModelAdmin):
    list_display = ['codigo']
    search_fields = ['codigo']
    actions = ['import_codigos']

    def qrcode_img(self, obj):
        if obj.qrcode:
            return mark_safe(f'<img src="{obj.qrcode.url}" height="100" />')
        return '-'

    qrcode_img.short_description = 'Código QR'


    def import_codigos(self, request, queryset):
        # Substitua 'caminho/para/planilha.xlsx' pelo caminho da sua planilha Excel
        file_path = 'app_core/planilha/planilha.xlsx'

        # Carrega a planilha Excel
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo, _ = row
            codigo_pulseira = CodigoPulseira(codigo=codigo)
            codigo_pulseira.save()

        self.message_user(request, f'{sheet.max_row - 1} códigos importados com sucesso.')

    import_codigos.short_description = 'Importar códigos da planilha Excel'
